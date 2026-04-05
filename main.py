from datetime import datetime
import json
import csv
import os
import matplotlib.pyplot as plt
from collections import defaultdict

def show_menu():
    print("\n1. Add Expenses")
    print("2. View Expenses")
    print("3. Total Expenses")
    print("4. Filter by Category")
    print("5. Filter by Date")
    print("6. Visualize expenses")
    print("7. Export report")
    print("8. Exit")

expenses=[]

def load_expenses():
    global expenses
    try:
        with open('expenses.json', 'r') as file:
            expenses=json.load(file)
            print(f"Loaded {len(expenses)} expenses") 
    except FileNotFoundError:
        expenses=[]
        print("No previous expenses found") 
load_expenses()


def save_expenses():
    print("Saving to:", os.path.abspath("expenses.json"))
    with open('expenses.json', 'w') as file:
        json.dump(expenses, file, indent=4)
    print(f"✓ Saved {len(expenses)} expenses")


def Add_Expenses():
    amount=float(input("Enter amount: "))
    category=input("Enter category: ")
    description=input("Enter description: ")

    date=datetime.now().strftime("%Y-%m-%d")

    expenses.append({
        'amount':amount,
        'category':category,
        'description':description,
        'date':date
        })
    
    print(f"Expenses list now has {len(expenses)} items")

    save_expenses()

    print("Expense Added!\n")
    


def View_Expenses():
    if not expenses:
        print("No Expenses Yet!\n")
        return
    
    for i,e in enumerate(expenses, start=1):
        print(f"{i}.{e.get('date', 'no date')} | ₹{e['amount']} | {e['category']} | {e['description']}")

    print()


def export_to_csv(expenses_to_export,filename):
    if not expenses_to_export:
        print("No data to export. ")
        return
    
    with open(filename, 'w', newline='') as file:
        writer=csv.writer(file)
        #header row
        writer.writerow(['Date','Amount','Category','Description'])

        for e in expenses_to_export:
            writer.writerow([
                e.get('date', 'No Date'),
                e['amount'],
                e['category'],
                e['description']
            ])

    print(f"Exported to {filename}\n")

def export_to_excel(expenses_to_export, filename):
    """Export expenses to Excel with formatting"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expenses"
        
        # Header styling
        headers = ['Date', 'Amount', 'Category', 'Description']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFFFF")  # Fixed: Added FF prefix
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row, e in enumerate(expenses_to_export, 2):
            ws.cell(row=row, column=1, value=e.get('date', 'No Date'))
            ws.cell(row=row, column=2, value=e['amount'])
            ws.cell(row=row, column=3, value=e['category'])
            ws.cell(row=row, column=4, value=e['description'])
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2
        
        wb.save(filename)
        print(f"✓ Exported to {filename}\n")
        
    except ImportError:
        print("Excel export requires openpyxl. Install with: pip install openpyxl")
        print("Falling back to CSV export...")
        export_to_csv(expenses_to_export, filename.replace('.xlsx', '.csv'))

def export_to_pdf(expenses_to_export, filename):
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import inch
        
        doc = SimpleDocTemplate(filename, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title = Paragraph("<b>Expense Report</b>", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))
        
        # Summary statistics
        total = sum(e['amount'] for e in expenses_to_export)
        avg = total / len(expenses_to_export) if expenses_to_export else 0
        
        summary = Paragraph(f"""
            <b>Summary:</b><br/>
            Total Expenses: ₹{total:,.2f}<br/>
            Number of Transactions: {len(expenses_to_export)}<br/>
            Average Transaction: ₹{avg:,.2f}<br/>
            Date Range: {expenses_to_export[0].get('date', 'N/A')} to {expenses_to_export[-1].get('date', 'N/A')}
        """, styles['Normal'])
        elements.append(summary)
        elements.append(Spacer(1, 0.3*inch))
        
        # Expenses table
        data = [['Date', 'Amount', 'Category', 'Description']]
        for e in expenses_to_export:
            data.append([
                e.get('date', 'No Date'),
                f"₹{e['amount']:.2f}",
                e['category'],
                e['description']
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        print(f"✓ PDF report generated: {filename}\n")
    except ImportError:
        print("PDF export requires reportlab. Install with: pip install reportlab")
        print("Falling back to CSV export...")
        export_to_csv(expenses_to_export, filename.replace('.pdf', '.csv'))


def visualize_expenses():
    """Show expense visualizations"""
    if not expenses:
        print("\nNo expenses to visualize.\n")
        return

    category_totals = defaultdict(float)
    date_totals = defaultdict(float)
    
    for e in expenses:
        category_totals[e['category']] += e['amount']
        date_totals[e.get('date', 'No Date')] += e['amount']
    
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))
    fig.suptitle('Expense Analysis', fontsize=16, fontweight='bold')
    
    categories = list(category_totals.keys())
    amounts = list(category_totals.values())
    colors_list = plt.cm.Set3(range(len(categories)))
    
    ax1.pie(amounts, labels=categories, autopct='%1.1f%%', colors=colors_list, startangle=90)
    ax1.set_title('Expenses by Category')
    
    dates = sorted(date_totals.keys())
    date_amounts = [date_totals[d] for d in dates]
    
    ax2.bar(range(len(dates)), date_amounts, color='steelblue')
    ax2.set_xlabel('Date')
    ax2.set_ylabel('Amount (₹)')
    ax2.set_title('Daily Spending')
    ax2.set_xticks(range(len(dates)))
    ax2.set_xticklabels(dates, rotation=45, ha='right')
    
    plt.tight_layout()
    plt.show()
    print("\n✓ Visualization displayed!\n")


def Total_Expenses():
    print("HELLO FROM TOTAL EXPENSES FUNCTION!")  # Test if function is called
    total = sum(e['amount'] for e in expenses)
    print(f"\nTotal Expenses: ₹{total}\n")


def Filter_by_Category():
    if not expenses:
        print("\nNo Expenses Yet.\n")
        return
    
    category=input("Enter category to filter: ").lower()
    found=False
    print()

    for i,e in enumerate(expenses, start=1):
        if e['category'].lower()==category:
            print(f"{i}.{e.get('date', 'no date')} | ₹{e['amount']} | {e['category']} | {e['description']}")
            found=True

    if not found:
        print("No expenses found in this category.")
    print()


def filter_by_date():
    if not expenses:
        print("\nNo expenses yet to filter.\n")
        return
    
    start_date=input("Enter start date (YYYY-MM-DD): ")
    end_date=input("Enter end date (YYYY-MM-DD): ")

    filtered=[]
    print()
    
    for e in expenses:
        expense_date=e.get('date')
        if expense_date and start_date <= expense_date <= end_date:
            filtered.append(e)
    
    if not filtered:
        print("No expenses found in this date range. \n")
        return
    
    for i,e in enumerate(filtered, start=1):
        print(f"{i}.{e.get('date')} | ₹{e['amount']} | {e['category']} | {e['description']}")
    
    print()

    export=input("Do you want to export these results to CSV? (y/n): ").lower()
    if export=='y':
        export_to_csv(filtered,"filtered_expenses.csv")


def export_menu(data_to_export=None):
    if data_to_export is None:
        data_to_export = expenses
    
    if not data_to_export:
        print("\nNo data to export.\n")
        return
    
    print("\nExport Options:")
    print("1. CSV (Excel-compatible)")
    print("2. Excel (.xlsx)")
    print("3. PDF Report")
    print("4. Cancel")
    
    choice = input("\nChoose export format: ")
    
    if choice == '1':
        filename = input("Enter filename (default: expenses.csv): ").strip() or "expenses.csv"
        if not filename.endswith('.csv'):
            filename += '.csv'
        export_to_csv(data_to_export, filename)
    
    elif choice == '2':
        filename = input("Enter filename (default: expenses.xlsx): ").strip() or "expenses.xlsx"
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        export_to_excel(data_to_export, filename)
    elif choice == '3':
        filename = input("Enter filename (default: expense_report.pdf): ").strip() or "expense_report.pdf"
        if not filename.endswith('.pdf'):
            filename += '.pdf'
        export_to_pdf(data_to_export, filename)
    
    elif choice == '4':
        print("Export cancelled.\n")
    
    else:
        print("Invalid choice.\n")


def main():

    while True:

        show_menu()

        choice=input("Choose an option:")
        if choice=='1':
            Add_Expenses()
        elif choice=='2':
            View_Expenses()
        elif choice=='3':
            Total_Expenses()
        elif choice=='4':
            Filter_by_Category()
        elif choice=='5':
            filter_by_date()
        elif choice=='6':
            visualize_expenses()
        elif choice=='7':
            export_menu()
        elif choice=='8':
            print("Exiting...")
            break
        else:
            print("Invalid Choice. Try Again.")

main()



